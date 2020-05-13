from data_processing1 import reader, counter
from openpyxl import load_workbook
import data_processing2 as dp
import os
import datetime
import random
from num2words import num2words

# данные из лр1 и лр2
in_min = 0.0
out_min = 2.0
sms = 2.0
in_min_bonus = 0
out_min_bonus = 20
sms_bonus = 0
customer_id = "915783624"

uslugi = []
plan = 0

in_min_am, out_min_am, sms_am = reader(customer_id)
cost_tel_in, cost_tel_out, cost_sms = counter(in_min_am, out_min_am, sms_am, in_min, out_min, sms, in_min_bonus, out_min_bonus, sms_bonus)

uslugi.append({'name': 'Входящие звонки', 'count': max(in_min_am - in_min_bonus, 0), 'ed': 'мин', 'price': in_min, "summ": cost_tel_in})
uslugi.append({'name': 'Исходящие звонки', 'count': max(out_min_am - out_min_bonus, 0), 'ed': 'мин', 'price': out_min, "summ": cost_tel_out})
uslugi.append({'name': 'СМС', 'count': max(sms_am - sms_bonus, 0), 'ed': 'шт', 'price': sms, "summ": cost_sms})

def mode_0():
    """Собирает необходимую для счета информацию
    """
    print("Подождите...")
    # Мб - мегабайты, Мбит - мегабиты (8*Мб)
    ip = "192.168.250.27"
    price = 1.0 # руб/Мб
    bonus = 0.0 # Мб
    file_nm = "ip_address.txt"

    if dp.check_file("nfcapd.202002251200"):
        mode_0()
    else:
        dp.translation()
        dp.ip_selection("result.txt", ip)
        traf = dp.traf_collection(file_nm)
        cost = dp.tariff_ip(traf, price, bonus) #байты, руб/Мб, Мб
        uslugi.append({'name': 'Интернет', 'count': traf, 'ed': 'байт', 'price': price, "summ": cost})
        # dp.diagram(file_nm)
        os.remove("result.txt")
        os.remove("ip_address.txt")

mode_0()

doc = load_workbook('template.xlsx') # открываем файл
ws = doc.active # проверяем какой лист активен
it = 23

ws.cell(column=2, row=2, value='ПАО "Яблоко", г. Сыктывкар')
inn = random.randint(1000000000, 9999999999)
kpp = random.randint(1000000000, 9999999999)
ws.cell(column=30, row=2, value=str(random.randint(1000000000, 9999999999))) #БИК 
ws.cell(column=30, row=3, value=str(random.randint(10000000000000000000, 99999999999999999999)))
ws.cell(column=30, row=5, value=str(random.randint(10000000000000000000, 99999999999999999999)))
ws.cell(column=2, row=6, value='ООО "Слон ест луну"')
ws.cell(column=7, row=14, value='ООО "Слон ест луну", ИНН: {}, КПП: {},\n192239, г.Санкт-Петербург, пер. Альпийский, 15/2'.format(inn, kpp))
ws.cell(column=7, row=17, value='ООО "Люся печет блины", ИНН: {}, КПП: {},\n197101, г.Санкт-Петербург, ул. Мира, 28'.format(random.randint(1000000000, 9999999999), random.randint(1000000000, 9999999999)))
ws.cell(column=7, row=20, value='№{} от 13.05.2020'.format(random.randint(1000000000, 9999999999)))
ws.cell(column=5, row=5, value=inn)
ws.cell(column=15, row=5, value=kpp)
ws.cell(column=2, row=10, value='Счет на оплату № {} от {} г.'.format(random.randint(1,10), str(datetime.datetime.now().strftime("%d.%m.%Y"))))

full_summ  = 0

for i in uslugi:
    ws.cell(column=4, row=it, value=i['name'])
    ws.cell(column=25, row=it, value=i['count'])
    ws.cell(column=29, row=it, value=i['ed'])
    ws.cell(column=32, row=it, value=i['price'])
    ws.cell(column=37, row=it, value=i['summ'])
    full_summ += i['summ']
    it += 1
ws.cell(column=38, row=28, value=str(round(full_summ, 2))+" руб.")
ws.cell(column=38, row=29, value=str(round(full_summ*0.2, 2))+" руб.")
ws.cell(column=38, row=30, value=str(round(full_summ*1.2, 2))+" руб.")
ws.cell(column=13, row=40, value='Ветрова О.П.')
ws.cell(column=36, row=40, value='Королева А.А.')
ws.cell(column=2, row=31, value='Всего наименований {}, на сумму {} руб.'.format(len(uslugi),round(full_summ*1.2, 2)))
ws.cell(column=2, row=32, value=num2words(round(full_summ*1.2), lang='ru') + ' рублей ' + str(int(round(full_summ*1.2, 2)*100%100)) + ' копеек')

doc.save('sample.xlsx')

os.system("libreoffice --headless --invisible --convert-to pdf result.pdf sample.xlsx >/dev/null 2>&1 ")
os.remove("sample.xlsx")
print("\nФайл 'sample.pdf' успешно создан")
input("\nНажмите Enter для выхода...")